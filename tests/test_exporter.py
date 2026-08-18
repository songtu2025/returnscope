from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from return_semantics.data import ReturnDataset
from return_semantics.exporter import export_results
from return_semantics.schemas import ValidatedClassification


def test_exporter_creates_expected_sheets(tmp_path: Path, taxonomy) -> None:
    classification_key = "APPAREL_TOO_SMALL\x1ftoo small"
    records = pd.DataFrame(
        [
            {
                "source_row": 2,
                "return-date": "2026-01-01",
                "order-id": "ORDER-1",
                "sku": "SKU-1",
                "asin": "ASIN-1",
                "reason": "APPAREL_TOO_SMALL",
                "comment_raw": "Too small",
                "comment_normalized": "Too small",
                "has_text_evidence": True,
                "classification_key": classification_key,
                "category_a": "水鞋",
                "category_b": "薄底水鞋",
            }
        ]
    )
    unique_comments = pd.DataFrame(
        [
            {
                "classification_key": classification_key,
                "reason": "APPAREL_TOO_SMALL",
                "comment_normalized": "Too small",
                "record_count": 1,
            }
        ]
    )
    dataset = ReturnDataset(
        records=records,
        unique_comments=unique_comments,
        mskus=frozenset({"SKU-1"}),
    )
    result = ValidatedClassification.model_validate(
        {
            "classification_key": classification_key,
            "semantic_units": [
                {
                    "subject": "PRODUCT",
                    "label_code": "FIT_TOO_SMALL",
                    "opinion": "尺码偏小",
                    "sentiment": "NEGATIVE",
                    "assertion": "AFFIRMED",
                    "part": "WHOLE_SHOE",
                    "evidence": "Too small",
                    "implicit": False,
                    "claim_relation": "NONE",
                    "claim_id": None,
                }
            ],
            "unknown_semantics": [],
            "problem_label_codes": ["FIT_TOO_SMALL"],
            "positive_label_codes": [],
            "primary_label_codes": ["FIT_TOO_SMALL"],
            "status": "AUTO_APPROVED",
            "review_reasons": [],
            "model_name": "test-model",
            "prompt_version": "test-prompt",
            "taxonomy_version": taxonomy.version,
        }
    )
    output_path = tmp_path / "result.xlsx"

    export_results(
        output_path,
        dataset,
        {classification_key: result},
        taxonomy,
    )

    workbook = load_workbook(output_path, read_only=True)
    assert workbook.sheetnames == [
        "分类明细",
        "语义单元",
        "人工复核",
        "未知语义",
        "标签统计",
    ]
    assert workbook["分类明细"].max_row == 2


def test_exporter_marks_missing_category_as_excluded(tmp_path: Path, taxonomy) -> None:
    classification_key = "SKU=SKU-2\x1fUNKNOWN\x1fnot configured"
    records = pd.DataFrame(
        [
            {
                "source_row": 2,
                "return-date": "2026-01-01",
                "order-id": "ORDER-2",
                "sku": "SKU-2",
                "asin": "ASIN-2",
                "reason": "UNKNOWN",
                "comment_raw": "Not configured",
                "comment_normalized": "Not configured",
                "has_text_evidence": True,
                "classification_key": classification_key,
                "category_a": "",
                "category_b": "",
            }
        ]
    )
    unique_comments = pd.DataFrame(
        [
            {
                "classification_key": classification_key,
                "reason": "UNKNOWN",
                "comment_normalized": "Not configured",
                "record_count": 1,
            }
        ]
    )
    dataset = ReturnDataset(
        records=records,
        unique_comments=unique_comments,
        mskus=frozenset({"SKU-2"}),
    )
    output_path = tmp_path / "excluded.xlsx"

    export_results(output_path, dataset, {}, taxonomy)

    detail = pd.read_excel(output_path, sheet_name="分类明细", dtype=str)
    assert detail.loc[0, "处理状态"] == "EXCLUDED_MISSING_CATEGORY"
