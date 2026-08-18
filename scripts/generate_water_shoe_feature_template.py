from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

SOURCE_HEADERS = {
    "return-date",
    "sku",
    "asin",
    "fnsku",
    "product-name",
}

SKU_HEADERS = [
    "canonical_sku",
    "asin",
    "reference_fnsku",
    "product_name_latest_observed",
    "product_name_version_count",
    "source_sku_aliases",
    "return_record_count_raw",
    "category_id",
    "product_family_id",
    "size_label",
    "size_system",
    "color",
    "package_configuration",
    "fact_source",
    "fact_updated_at",
    "width_design",
    "toe_box_design",
    "upper_material",
    "outsole_material",
    "insole_material",
    "sole_structure",
    "drainage_structure",
    "closure_style",
    "fact_notes",
]

SKU_HEADER_COMMENTS = {
    "canonical_sku": "程序生成的规范 SKU，也是产品特征关联主键，请勿修改。",
    "asin": "当前退货报告中与规范 SKU 唯一对应的 ASIN，请勿修改。",
    "reference_fnsku": "排除 FNSKU=ASIN 等占位值后的参考 FNSKU，请勿修改。",
    "product_name_latest_observed": "退货报告中最近观察到的历史标题，仅用于识别产品，不能作为客观特征。",
    "product_name_version_count": "该 SKU 在当前报告中出现过的标题版本数。",
    "source_sku_aliases": "Amazon.Found.<ASIN> 等原始占位 SKU，供追溯使用。",
    "return_record_count_raw": "当前报告中关联到该 SKU 的原始记录数，包含完全重复记录。",
    "category_id": "首版固定为 water_shoes。",
    "product_family_id": "必填。相同材料、结构和 Listing 声明的一组 SKU 使用同一个产品型号编号。",
    "size_label": "必填。Listing 上展示给买家的完整尺码文字，不要只从 SKU 猜测。",
    "size_system": "必填。选择 US_UNISEX、US_WOMEN、US_MEN、US_KIDS、EU、UK 或 OTHER。",
    "color": "必填。Listing 上展示给买家的颜色名称。",
    "package_configuration": "必填。例如 1 pair / 2 shoes。",
    "fact_source": "必填。客观特征依据，例如 product_spec 或 sample_measurement。",
    "fact_updated_at": "必填。当前客观特征最后确认日期。",
    "width_design": "可选。仅填写可验证的楦型宽度设计，不填写 wide fit 等营销结论。",
    "toe_box_design": "可选。仅填写可验证的鞋头结构宽度。",
    "upper_material": "可选。鞋面客观材质。",
    "outsole_material": "可选。外底客观材质。",
    "insole_material": "可选。鞋垫客观材质。",
    "sole_structure": "可选。鞋底厚度、纹路或分层结构等客观描述。",
    "drainage_structure": "可选。排水孔、导流槽等客观结构；不要在这里填写 quick dry。",
    "closure_style": "可选。例如 slip_on、lace、hook_loop。",
    "fact_notes": "可选。只记录产品事实补充，不记录客户投诉。",
}

CLAIM_HEADERS = [
    "product_family_id",
    "canonical_sku",
    "claim_type",
    "claim_text",
    "source_location",
    "effective_from",
    "effective_to",
    "captured_at",
    "claim_notes",
]

CLAIM_HEADER_COMMENTS = {
    "product_family_id": "必填。必须与 SKU客观特征 工作表中的 product_family_id 一致。",
    "canonical_sku": "可空。为空表示该声明适用于整个产品型号；有值表示只适用于该 SKU。",
    "claim_type": "必填。声明类型使用下拉枚举。",
    "claim_text": "必填。保留 Listing 中的原始英文声明，不要改写。",
    "source_location": "必填。声明出现的位置，例如 title、bullet、A+、size_chart。",
    "effective_from": "已知时填写声明生效日期；未知可空，但不能据此判断历史 Listing 责任。",
    "effective_to": "当前仍有效可空；历史声明填写失效日期。",
    "captured_at": "必填。你采集这条声明的日期。",
    "claim_notes": "可选。补充声明版本或适用范围。",
}

PREFILLED_FILL = PatternFill("solid", fgColor="D9EAF7")
REQUIRED_FILL = PatternFill("solid", fgColor="FFF2CC")
OPTIONAL_FILL = PatternFill("solid", fgColor="E2F0D9")
HEADER_FONT = Font(bold=True, color="1F1F1F")
MISSING_FILL = PatternFill("solid", fgColor="FCE8E6")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="生成涉水鞋 SKU 产品特征模板")
    parser.add_argument(
        "--source",
        type=Path,
        default=Path("input_data") / "SEEKWAY_US_.csv",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("input_data") / "涉水鞋_SKU产品特征模板.xlsx",
    )
    return parser.parse_args()


def is_placeholder_sku(sku: str) -> bool:
    return sku.startswith("Amazon.Found.")


def build_asin_sku_map(source_path: Path) -> dict[str, str]:
    asin_to_skus: dict[str, set[str]] = defaultdict(set)

    with source_path.open("r", encoding="cp1252", newline="") as source_file:
        reader = csv.DictReader(source_file)
        missing_headers = SOURCE_HEADERS.difference(reader.fieldnames or [])
        if missing_headers:
            missing_text = ", ".join(sorted(missing_headers))
            raise ValueError(f"退货报告缺少字段: {missing_text}")

        for row in reader:
            sku = row["sku"].strip()
            asin = row["asin"].strip()
            if not is_placeholder_sku(sku):
                asin_to_skus[asin].add(sku)

    invalid = {
        asin: sorted(skus) for asin, skus in asin_to_skus.items() if len(skus) != 1
    }
    if invalid:
        raise ValueError(f"ASIN 无法唯一映射规范 SKU: {invalid}")

    return {asin: next(iter(skus)) for asin, skus in asin_to_skus.items()}


def aggregate_skus(
    source_path: Path,
    asin_sku_map: dict[str, str],
) -> list[dict[str, object]]:
    aggregates: dict[str, dict[str, object]] = {}

    with source_path.open("r", encoding="cp1252", newline="") as source_file:
        reader = csv.DictReader(source_file)
        for row in reader:
            raw_sku = row["sku"].strip()
            asin = row["asin"].strip()
            canonical_sku = asin_sku_map.get(asin)
            if canonical_sku is None:
                raise ValueError(f"ASIN 没有规范 SKU: {asin}")
            if not is_placeholder_sku(raw_sku) and raw_sku != canonical_sku:
                raise ValueError(f"SKU 与 ASIN 映射冲突: {raw_sku}, {asin}")

            item = aggregates.setdefault(
                canonical_sku,
                {
                    "canonical_sku": canonical_sku,
                    "asins": set(),
                    "fnskus": set(),
                    "titles": set(),
                    "aliases": set(),
                    "latest_date": "",
                    "latest_title": "",
                    "record_count": 0,
                },
            )
            item["asins"].add(asin)
            item["record_count"] += 1

            if is_placeholder_sku(raw_sku):
                expected_alias = f"Amazon.Found.{asin}"
                if raw_sku != expected_alias:
                    raise ValueError(f"占位 SKU 与 ASIN 不一致: {raw_sku}, {asin}")
                item["aliases"].add(raw_sku)
                continue

            fnsku = row["fnsku"].strip()
            if fnsku and fnsku != asin:
                item["fnskus"].add(fnsku)

            title = row["product-name"].strip()
            if title:
                item["titles"].add(title)
                return_date = row["return-date"].strip()
                if return_date >= item["latest_date"]:
                    item["latest_date"] = return_date
                    item["latest_title"] = title

    output_rows = []
    for canonical_sku, item in aggregates.items():
        if len(item["asins"]) != 1:
            raise ValueError(f"规范 SKU 对应多个 ASIN: {canonical_sku}")
        if len(item["fnskus"]) != 1:
            raise ValueError(f"规范 SKU 无法唯一确定参考 FNSKU: {canonical_sku}")
        if not item["latest_title"]:
            raise ValueError(f"规范 SKU 缺少可识别标题: {canonical_sku}")

        output_rows.append(
            {
                "canonical_sku": canonical_sku,
                "asin": next(iter(item["asins"])),
                "reference_fnsku": next(iter(item["fnskus"])),
                "product_name_latest_observed": item["latest_title"],
                "product_name_version_count": len(item["titles"]),
                "source_sku_aliases": " | ".join(sorted(item["aliases"])),
                "return_record_count_raw": item["record_count"],
                "category_id": "water_shoes",
            }
        )

    return sorted(output_rows, key=lambda row: str(row["canonical_sku"]).lower())


def add_list_validation(
    sheet,
    cell_range: str,
    values: list[str],
) -> None:
    formula = '"' + ",".join(values) + '"'
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.error = "请从下拉列表中选择有效值"
    validation.errorTitle = "无效值"
    validation.prompt = "请选择下拉列表中的值；未知可留空"
    validation.promptTitle = "填写提示"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    sheet.add_data_validation(validation)
    validation.add(cell_range)


def style_sheet(
    sheet,
    headers: list[str],
    comments: dict[str, str],
    prefilled_count: int,
    required_count: int,
) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 32

    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index)
        cell.value = header
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.comment = Comment(comments[header], "Codex")

        if column_index <= prefilled_count:
            cell.fill = PREFILLED_FILL
        elif column_index <= prefilled_count + required_count:
            cell.fill = REQUIRED_FILL
        else:
            cell.fill = OPTIONAL_FILL

    widths = {
        "canonical_sku": 34,
        "asin": 16,
        "reference_fnsku": 18,
        "product_name_latest_observed": 72,
        "product_name_version_count": 18,
        "source_sku_aliases": 30,
        "return_record_count_raw": 20,
        "category_id": 16,
        "product_family_id": 22,
        "size_label": 22,
        "size_system": 16,
        "color": 20,
        "package_configuration": 22,
        "fact_source": 22,
        "fact_updated_at": 18,
        "width_design": 18,
        "toe_box_design": 18,
        "upper_material": 24,
        "outsole_material": 24,
        "insole_material": 24,
        "sole_structure": 32,
        "drainage_structure": 32,
        "closure_style": 20,
        "fact_notes": 40,
        "claim_type": 20,
        "claim_text": 60,
        "source_location": 20,
        "effective_from": 18,
        "effective_to": 18,
        "captured_at": 18,
        "claim_notes": 40,
    }
    for index, header in enumerate(headers, start=1):
        letter = sheet.cell(row=1, column=index).column_letter
        sheet.column_dimensions[letter].width = widths.get(header, 20)


def build_workbook(rows: list[dict[str, object]], output_path: Path) -> None:
    workbook = Workbook()
    sku_sheet = workbook.active
    sku_sheet.title = "SKU客观特征"
    style_sheet(
        sku_sheet,
        SKU_HEADERS,
        SKU_HEADER_COMMENTS,
        prefilled_count=8,
        required_count=7,
    )

    for row_index, row in enumerate(rows, start=2):
        for column_index, header in enumerate(SKU_HEADERS, start=1):
            sku_sheet.cell(row=row_index, column=column_index).value = row.get(header)
        sku_sheet.cell(row=row_index, column=15).number_format = "yyyy-mm-dd"

    last_sku_row = len(rows) + 1
    sku_table = Table(displayName="SkuObjectiveFacts", ref=f"A1:X{last_sku_row}")
    sku_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sku_sheet.add_table(sku_table)

    add_list_validation(
        sku_sheet,
        f"K2:K{last_sku_row}",
        ["US_UNISEX", "US_WOMEN", "US_MEN", "US_KIDS", "EU", "UK", "OTHER"],
    )
    add_list_validation(
        sku_sheet,
        f"N2:N{last_sku_row}",
        ["product_spec", "sample_measurement", "supplier_spec", "lab_test", "mixed"],
    )
    add_list_validation(
        sku_sheet,
        f"P2:Q{last_sku_row}",
        ["narrow", "standard", "wide", "unknown"],
    )
    sku_sheet.conditional_formatting.add(
        f"I2:O{last_sku_row}",
        FormulaRule(formula=['I2=""'], fill=MISSING_FILL),
    )

    claim_sheet = workbook.create_sheet("Listing声明")
    style_sheet(
        claim_sheet,
        CLAIM_HEADERS,
        CLAIM_HEADER_COMMENTS,
        prefilled_count=0,
        required_count=8,
    )
    claim_sheet.append([None] * len(CLAIM_HEADERS))
    for column_index in (6, 7, 8):
        claim_sheet.cell(row=2, column=column_index).number_format = "yyyy-mm-dd"

    claim_table = Table(displayName="ListingClaims", ref="A1:I2")
    claim_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    claim_sheet.add_table(claim_table)
    add_list_validation(
        claim_sheet,
        "C2:C10000",
        [
            "size_fit",
            "width_fit",
            "material",
            "quick_dry",
            "drainage",
            "slip_resistance",
            "breathability",
            "protection",
            "intended_use",
            "package_quantity",
            "other",
        ],
    )
    add_list_validation(
        claim_sheet,
        "E2:E10000",
        ["title", "bullet", "A+", "size_chart", "package", "manual", "other"],
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def verify_workbook(output_path: Path, expected_rows: int) -> None:
    workbook = load_workbook(output_path, read_only=False, data_only=False)
    if workbook.sheetnames != ["SKU客观特征", "Listing声明"]:
        raise ValueError(f"工作表不符合预期: {workbook.sheetnames}")

    sku_sheet = workbook["SKU客观特征"]
    if sku_sheet.max_row != expected_rows + 1:
        raise ValueError("SKU 模板行数与生成结果不一致")

    headers = [cell.value for cell in sku_sheet[1]]
    if headers != SKU_HEADERS:
        raise ValueError("SKU 模板字段与契约不一致")

    sku_values = [
        sku_sheet.cell(row=row, column=1).value
        for row in range(2, sku_sheet.max_row + 1)
    ]
    if len(sku_values) != len(set(sku_values)):
        raise ValueError("SKU 模板中存在重复 canonical_sku")


def main() -> None:
    args = parse_args()
    asin_sku_map = build_asin_sku_map(args.source)
    rows = aggregate_skus(args.source, asin_sku_map)
    build_workbook(rows, args.output)
    verify_workbook(args.output, expected_rows=len(rows))
    print(f"已生成: {args.output}")
    print(f"规范 SKU 数: {len(rows)}")


if __name__ == "__main__":
    main()
